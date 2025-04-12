from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from io import BytesIO
from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm
import random
import xlwt


# ---------- Login/Register Views ----------

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('library-dashboard')  # Redirect to the library dashboard after login
    else:
        form = UserCreationForm()
    
    return render(request, 'registration/register.html', {'form': form})

# ---------- DashBoard View ----------


def library_dashboard(request):
    # List of book-related quotes
    quotes = [
        "A room without books is like a body without a soul. – Marcus Tullius Cicero",
        "The only thing you absolutely have to know is the location of the library. – Albert Einstein",
        "Books are a uniquely portable magic. – Stephen King",
        "There is no friend as loyal as a book. – Ernest Hemingway",
        "A book is a dream that you hold in your hand. – Neil Gaiman",
        "So many books, so little time. – Frank Zappa",
        "The more that you read, the more things you will know. The more that you learn, the more places you'll go. – Dr. Seuss",
        "Books are the mirrors of the soul. – Virginia Woolf"
    ]
    
    # Select a random quote
    random_quote = random.choice(quotes)

    # Pass the quote to the template
    return render(request, 'manager/dashboard.html', {'random_quote': random_quote})

# ---------- LIST VIEWS (with Pagination) ----------

class AuthorListView(ListView):
    model = Author
    template_name = 'manager/author_list.html'
    context_object_name = 'authors'
    paginate_by = 5
    ordering = ['name']  


class BookListView(ListView):
    model = Book
    template_name = 'manager/book_list.html'
    context_object_name = 'books'
    paginate_by = 5
    ordering = ['title']  


class BorrowRecordListView(ListView):
    model = BorrowRecord
    template_name = 'manager/borrow_list.html'
    context_object_name = 'borrow_records'
    paginate_by = 5
    ordering = ['-borrow_date']  


# ---------- CREATE VIEWS ----------

class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'manager/author_form.html'
    success_url = reverse_lazy('author-list')

    def form_valid(self, form):
        messages.success(self.request, "Author added successfully.")
        return super().form_valid(form)


class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'manager/book_form.html'
    success_url = reverse_lazy('book-list')

    def form_valid(self, form):
        messages.success(self.request, "Book added successfully.")
        return super().form_valid(form)


class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'manager/borrow_form.html'
    success_url = reverse_lazy('borrow-list')

    def form_valid(self, form):
        messages.success(self.request, "Borrow record added successfully.")
        return super().form_valid(form)


# ---------- UPDATE VIEWS ----------

class AuthorUpdateView(UpdateView):
    model = Author
    form_class = AuthorForm
    template_name = 'manager/edit_author.html'
    success_url = reverse_lazy('author-list')

    def form_valid(self, form):
        messages.success(self.request, "Author updated successfully.")
        return super().form_valid(form)


class BookUpdateView(UpdateView):
    model = Book
    form_class = BookForm
    template_name = 'manager/edit_book.html'
    success_url = reverse_lazy('book-list')

    def form_valid(self, form):
        messages.success(self.request, "Book updated successfully.")
        return super().form_valid(form)


class BorrowRecordUpdateView(UpdateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'manager/edit_borrow.html'
    success_url = reverse_lazy('borrow-list')

    def form_valid(self, form):
        messages.success(self.request, "Borrow record updated successfully.")
        return super().form_valid(form)


# ---------- DELETE VIEWS ----------

class AuthorDeleteView(DeleteView):
    model = Author
    success_url = reverse_lazy('author-list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Author deleted successfully.")
        return super().delete(request, *args, **kwargs)


class BookDeleteView(DeleteView):
    model = Book
    success_url = reverse_lazy('book-list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Book deleted successfully.")
        return super().delete(request, *args, **kwargs)


class BorrowRecordDeleteView(DeleteView):
    model = BorrowRecord
    success_url = reverse_lazy('borrow-list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Borrow record deleted successfully.")
        return super().delete(request, *args, **kwargs)


# ---------- EXPORT TO EXCEL ----------
def export_authors_excel(request):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet('Authors')

    authors = Author.objects.all()

    sheet.write(0, 0, 'ID')
    sheet.write(0, 1, 'Name')

    for i, author in enumerate(authors, start=1):
        sheet.write(i, 0, author.id)
        sheet.write(i, 1, author.name)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=authors.xls'
    wb.save(response)
    return response


def export_books_excel(request):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet('Books')

    books = Book.objects.select_related('author').all()

    sheet.write(0, 0, 'ID')
    sheet.write(0, 1, 'Title')
    sheet.write(0, 2, 'Author')
    sheet.write(0, 3, 'Published Date')

    for i, book in enumerate(books, start=1):
        sheet.write(i, 0, book.id)
        sheet.write(i, 1, book.title)
        sheet.write(i, 2, book.author.name if book.author else '')
        sheet.write(i, 3, str(book.published_date))

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=books.xls'
    wb.save(response)
    return response


def export_borrowrecords_excel(request):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet('Borrow Records')

    borrowrecords = BorrowRecord.objects.select_related('book').all()

    # Header row
    sheet.write(0, 0, 'ID')
    sheet.write(0, 1, 'Book Title')
    sheet.write(0, 2, 'User Name')
    sheet.write(0, 3, 'Borrow Date')
    sheet.write(0, 4, 'Return Date')

    # Data rows
    for i, record in enumerate(borrowrecords, start=1):
        sheet.write(i, 0, record.id)
        sheet.write(i, 1, record.book.title if record.book else '')
        sheet.write(i, 2, record.user_name)
        sheet.write(i, 3, str(record.borrow_date))
        sheet.write(i, 4, str(record.return_date))

    # Return response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=borrow_records.xls'
    wb.save(response)
    return response